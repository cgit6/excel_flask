import os
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView
from openpyxl.worksheet.properties import PageSetupProperties
from datetime import datetime
import re
import shutil
import time
import sys
import platform
import subprocess
import unicodedata
import pandas as pd
import json

# 檢測當前操作系統
CURRENT_OS = platform.system()  # 返回 'Windows', 'Linux', 或 'Darwin' (macOS)

# 在Windows下嘗試導入win32com庫
if CURRENT_OS == 'Windows':
    try:
        import win32com.client
        from pywintypes import com_error
        win32com_available = True
    except ImportError:
        win32com_available = False
        print("警告：Windows環境下未安裝win32com庫，部分PDF功能可能受限")
        print("可以使用 pip install pywin32 安裝此庫")
else:
    win32com_available = False

# 獲取字符的視覺寬度
def get_visual_width(text):
    width = 0
    for char in text:
        # 獲取字符的東亞寬度屬性
        if unicodedata.east_asian_width(char) in ['F', 'W', 'A']:
            width += 2  # 全角字符寬度為2
        else:
            width += 1  # 半角字符寬度為1
    return width

# 根據視覺寬度進行自動換行
def wrap_by_visual_width(text, max_width):
    if not text:
        return []
        
    result = []
    current_line = ""
    current_width = 0
    
    # 分割文本（可以按空格分割）
    words = text.split()
    
    for word in words:
        word_width = get_visual_width(word)
        
        # 如果單詞過長，需要拆分字符
        if word_width > max_width:
            if current_line:  # 先處理當前行
                result.append(current_line)
                current_line = ""
                current_width = 0
            
            # 按字符拆分長單詞
            temp_word = ""
            temp_width = 0
            for char in word:
                char_width = get_visual_width(char)
                if temp_width + char_width <= max_width:
                    temp_word += char
                    temp_width += char_width
                else:
                    result.append(temp_word)
                    temp_word = char
                    temp_width = char_width
            
            if temp_word:  # 添加最後一部分
                current_line = temp_word
                current_width = temp_width
        else:
            # 檢查是否需要換行
            if current_width + 1 + word_width > max_width and current_line:  # +1 是為了空格
                result.append(current_line)
                current_line = word
                current_width = word_width
            else:
                if current_line:
                    current_line += " " + word
                    current_width += 1 + word_width  # +1 是為了空格
                else:
                    current_line = word
                    current_width = word_width
    
    if current_line:  # 添加最後一行
        result.append(current_line)
    
    return result

def export_excel_to_pdf(excel_path, pdf_path):
    """
    將Excel文件導出為PDF格式，根據操作系統使用不同方法
    
    參數:
    excel_path (str): Excel文件路徑
    pdf_path (str): 輸出PDF文件路徑
    
    返回:
    bool: 成功返回True，失敗返回False
    """
    # 獲取文件的絕對路徑
    excel_path = os.path.abspath(excel_path)
    pdf_path = os.path.abspath(pdf_path)
    
    # 確保輸出目錄存在
    pdf_dir = os.path.dirname(pdf_path)
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir)
        
    print(f"正在將Excel轉換為PDF: {excel_path} -> {pdf_path}")
    
    # 根據操作系統選擇不同的轉換方法
    if CURRENT_OS == 'Windows' and win32com_available:
        return export_excel_to_pdf_windows(excel_path, pdf_path)
    else:
        return export_excel_to_pdf_linux(excel_path, pdf_path)

def export_excel_to_pdf_windows(excel_path, pdf_path):
    """Windows環境下使用win32com將Excel轉換為PDF"""
    try:
        # 先初始化COM環境
        import pythoncom
        pythoncom.CoInitialize()
        
        try:
            # 初始化Excel應用程序
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # 不顯示Excel窗口
            
            try:
                # 打開工作簿
                wb = excel.Workbooks.Open(excel_path)
                
                # 設置打印選項
                ws = wb.ActiveSheet
                ws.PageSetup.Zoom = False  # 禁用縮放
                ws.PageSetup.FitToPagesWide = 1  # 適應1頁寬
                ws.PageSetup.FitToPagesTall = False  # 高度不限制
                
                # 轉換為PDF (xlTypePDF = 0)
                wb.ExportAsFixedFormat(0, pdf_path)
                
                # 關閉工作簿而不保存
                wb.Close(False)
                print(f"PDF導出成功: {pdf_path}")
                return True
                
            except Exception as e:
                print(f"導出PDF時出錯: {str(e)}")
                return False
                
            finally:
                # 退出Excel應用程序
                excel.Quit()
        
        except Exception as e:
            print(f"創建Excel應用程序時出錯: {str(e)}")
            return False
            
        finally:
            # 釋放COM資源
            pythoncom.CoUninitialize()
        
    except Exception as e:
        print(f"COM環境初始化失敗: {str(e)}")
        return False

def export_excel_to_pdf_linux(excel_path, pdf_path):
    """Linux環境下使用LibreOffice將Excel轉換為PDF"""
    try:
        # 檢查LibreOffice是否可用
        try:
            # 檢查soffice命令是否存在
            subprocess.run(['which', 'soffice'], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except subprocess.CalledProcessError:
            print("錯誤: 在Linux環境下需要安裝LibreOffice")
            print("請使用以下命令安裝: sudo apt-get install libreoffice")
            return False
            
        # 構建轉換命令
        cmd = [
            'soffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', os.path.dirname(pdf_path),
            excel_path
        ]
        
        # 執行命令
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        # 檢查結果
        if result.returncode == 0:
            # 由於LibreOffice會使用原始文件名，我們可能需要重命名輸出文件
            original_pdf_name = os.path.basename(excel_path).replace('.xlsx', '.pdf')
            original_pdf_path = os.path.join(os.path.dirname(pdf_path), original_pdf_name)
            
            # 如果生成的PDF與目標路徑不同，則重命名
            if original_pdf_path != pdf_path and os.path.exists(original_pdf_path):
                os.rename(original_pdf_path, pdf_path)
                
            print(f"PDF導出成功: {pdf_path}")
            return True
        else:
            print(f"LibreOffice轉換失敗: {result.stderr.decode()}")
            return False
            
    except Exception as e:
        print(f"Linux環境下PDF導出過程中發生異常: {str(e)}")
        return False

# json 轉成 excel
def json_to_excel(json_data, temp_excel_path, save_original=False):
    """
    將JSON資料轉換為Excel檔案
    
    參數:
    json_data (list/str): JSON資料或JSON字串
    temp_excel_path (str): 臨時Excel檔案儲存路徑
    save_original (bool): 是否同時保存原始轉換結果
    
    返回:
    bool: 成功返回True，失敗返回False
    """
    try:
        # 如果輸入是字串，嘗試解析為JSON
        if isinstance(json_data, str):
            json_data = json.loads(json_data)
        
        print(">>> json_to_excel parsed json_data:\n", json_data) # 打印 json_data
        
        # 檢查資料是否為列表
        if not isinstance(json_data, list):
            print("錯誤:輸入的JSON資料不是列表格式")
            return False
        
        
        # 檢查資料是否為空
        if len(json_data) == 0:
            print("錯誤:輸入的JSON資料為空")
            return False
        
        # 檢查每個訂單項是否包含必要欄位
        for i, order in enumerate(json_data):
            if "name" not in order:
                print(f"錯誤:第{i+1}項資料缺少'name'欄位")
                return False
            if "products" not in order:
                print(f"錯誤:第{i+1}項資料缺少'products'欄位")
                return False
            if not isinstance(order["products"], list):
                print(f"錯誤:第{i+1}項的'products'欄位不是列表格式")
                return False
            for j, product in enumerate(order["products"]):
                if "productName" not in product:
                    print(f"錯誤:第{i+1}項的第{j+1}個產品缺少'productName'欄位")
                    return False
                if "amount" not in product:
                    print(f"錯誤:第{i+1}項的第{j+1}個產品缺少'amount'欄位")
                    return False
        
        # 創建Excel工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 設置列標題
        ws.cell(row=1, column=1).value = "姓名"
        ws.cell(row=1, column=2).value = "商品"
        
        # 填充資料
        row_num = 2
        for order in json_data:
            # 將一個客戶的所有產品資訊合併成一個文本，用換行符分隔
            products_text = []
            for product in order["products"]:
                products_text.append(f"{product['productName']} x {product['amount']}")
            
            # 一個客戶只佔一行，所有產品放在同一個單元格中並用換行符分隔
            ws.cell(row=row_num, column=1).value = order["name"]
            ws.cell(row=row_num, column=2).value = "\n".join(products_text)
            row_num += 1
        
        # 保存原始轉換結果
        if save_original:
            original_path = os.path.join(os.path.dirname(temp_excel_path), "original_conversion.xlsx")
            wb.save(original_path)
            print(f"已保存原始轉換結果：{original_path}")
        
        # 保存Excel檔案
        wb.save(temp_excel_path)
        print(f"已將JSON資料轉換為Excel檔案：{temp_excel_path}")
        return True
    
    except Exception as e:
        print(f"JSON轉Excel過程中出錯：{str(e)}")
        import traceback
        print(traceback.format_exc())
        return False

def format_shipping_document(input_data, output_file, export_pdf=False, save_intermediate=False):
    """
    將Excel檔案或JSON資料按照指定格式進行處理，模擬VBA巨集'出貨單整理'的功能
    
    參數:
    input_data (str/list):輸入檔案路徑或JSON資料
    output_file (str):輸出檔案路徑
    export_pdf (bool):是否同時導出PDF檔案
    save_intermediate (bool):是否保存中間處理結果
    """
    try:
        print(f"檢測到操作系統: {CURRENT_OS}")
        
        # 確保輸出目錄存在
        output_dir = os.path.dirname(output_file) # 輸出目錄
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        # 如果輸出檔案已存在，嘗試先創建一個臨時輸出檔案
        if os.path.exists(output_file):
            base_name = os.path.splitext(output_file)[0] # 輸出檔案名稱
            ext = os.path.splitext(output_file)[1] # 輸出檔案副檔名
            temp_output = f"{base_name}_{int(time.time())}{ext}" # 臨時輸出檔案名稱
        else:
            temp_output = output_file
            
        # 定義實際使用的輸出檔案
        final_excel_path = output_file
            
        # 判斷輸入是檔案路徑還是JSON資料
        is_json_data = isinstance(input_data, (list, dict)) or (
            isinstance(input_data, str) and not os.path.exists(input_data) and (
                input_data.startswith('[') or input_data.startswith('{')
            )
        )
    
        # 如果輸入是JSON資料()
        if is_json_data:
            print("檢測到JSON資料輸入")
            # 創建臨時Excel檔案
            temp_excel = os.path.join(os.path.dirname(output_file), f"temp_{int(time.time())}.xlsx")
            
            # 將JSON轉換為Excel
            if not json_to_excel(input_data, temp_excel, save_original=save_intermediate):
                print("JSON轉Excel失敗")
                return False
                
            # 設定輸入檔案為臨時Excel檔案
            input_file = temp_excel
            
            # 加上黑框線，粗體字
            print(f"已將JSON資料轉換為臨時Excel檔案: {input_file}")
        else:
            # 輸入為檔案路徑
            input_file = input_data
            print(f"正在處理檔案: {input_file}")
            
            # 檢查輸入檔案是否存在
            if not os.path.exists(input_file):
                print(f"錯誤：輸入檔案 {input_file} 不存在!")
                return False
        
        # 載入工作簿
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active 
        
        # 如果是JSON資料輸入，為所有單元格添加黑框線和粗體字
        if is_json_data:
            # 定義黑色邊框樣式
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # 遍歷所有已使用的單元格，添加黑框線和粗體字
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    # 添加黑框線
                    cell.border = thin_border
                    # 確保文字是粗體
                    if cell.font:
                        new_font = Font(
                            name=cell.font.name if cell.font.name else "微軟正黑體",
                            size=cell.font.size if cell.font.size else 12,
                            bold=True
                        )
                        cell.font = new_font
                    else:
                        cell.font = Font(name="微軟正黑體", size=12, bold=True)
            
            print("已為JSON輸入資料添加黑框線和粗體字")
        
        # 自动调整A和B列的列宽 - 对应VBA的Columns("A:B").EntireColumn.AutoFit
        # openpyxl不直接支持自动调整列宽，我们使用固定寬度
        # A列(第1列，價格欄位)設置固定寬度為13.14
        ws.column_dimensions[get_column_letter(1)].width = 13.43
        
        # B列(第2列，商品欄位)設置固定寬度為124.14
        ws.column_dimensions[get_column_letter(2)].width = 124.43

        # 如果需要保存中間處理結果，先保存一份原始檔案
        if save_intermediate:
            intermediate_path = os.path.join(os.path.dirname(output_file), "original_input.xlsx")
            wb.save(intermediate_path)
            print(f"已保存原始從 format_shipping_document 輸入的檔案：{intermediate_path}")
        
        # 設置A1:Y100範圍的單元格格式 - 對應VBA的Range("A1:Y100").Select部分
        for row in range(1, 101):
            for col in range(1, 26):  # A到Y是1到25
                cell = ws.cell(row=row, column=col)
                # 設置垂直對齊和自動換行
                cell.alignment = Alignment(
                    vertical='center',
                    wrap_text=True
                )
                # 設置字體
                cell.font = Font(
                    name="微軟正黑體",
                    size=12,
                    bold=True
                )
                # 設置水平對齊
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='center',
                    wrap_text=True
                )
        
        if not is_json_data:
            # 如果傳入的資料是 excel 檔案則執行 for_excel
            # for_excel:删除B和C列 - 对应VBA的Columns("B:C").Select和Selection.Delete
            ws.delete_cols(2, 2)
                # 删除C:Y列 - 对应VBA的最后部分Columns("C:Y").Select和Selection.Delete
            if ws.max_column > 2:
                ws.delete_cols(3, ws.max_column - 2)  # 删除从第3列开始的所有列        
            
            # for_excel:替换B列中的逗号为换行符 - 对应VBA的Selection.Replace
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=2)  # B列现在是第2列
                if cell.value:
                    # 先按逗號分割文本
                    text = str(cell.value)
                    segments = text.split(',')
                    result_lines = []
                    
                    # 處理每個分段
                    for segment in segments:
                        segment = segment.strip()  # 去除首尾空白
                        
                        # # 檢查分段的視覺寬度
                        # if get_visual_width(segment) > 64:
                        #     # 如果超過64，則按視覺寬度進行換行
                        #     wrapped = wrap_by_visual_width(segment, 64)
                        #     result_lines.extend(wrapped)
                        # else:
                        #     # 不超過則直接添加
                        #     result_lines.append(segment)

                        result_lines.append(segment)
                    
                    # 將處理後的文本合併為以換行符分隔的字符串
                    cell.value = '\n'.join(result_lines)
                    print(f"行 {row} 商品文本處理後包含 {len(result_lines)} 行")
        


            # # 保存輸出檔案
            # if save_intermediate:
            #     wb.save(os.path.join(os.path.dirname(output_file), "before_final_formatting.xlsx"))
            #     print(f"已保存最終格式化前的結果")
        
        # 對檔案的樣式進行處理
        # 设置B列字体大小为20 - 对应VBA的With Selection.Font .Size = 20
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.font:
                new_font = Font(
                    name=cell.font.name,
                    size=20,
                    bold=cell.font.bold
                )
                cell.font = new_font
            else:
                cell.font = Font(size=20, bold=True)
        
        # 設置行高 = 單元格內文本行數 + 1
        for row in range(1, ws.max_row + 1):
            # 獲取A列和B列的單元格
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            
            # 默認行數為1（即使單元格為空）
            line_count_a = 1
            line_count_b = 1
            
            # 計算A列的文本行數
            if cell_a.value:
                text_a = str(cell_a.value)
                line_count_a = text_a.count('\n') + 1
                
            # 計算B列的文本行數
            if cell_b.value:
                text_b = str(cell_b.value)
                line_count_b = text_b.count('\n') + 1
            
            # 使用A列和B列中較大的行數
            line_count = max(line_count_a, line_count_b)
            
            # 設置行高 = 行數 + 1（以磅為單位）
            # 確保至少有2行的高度
            row_height = max(2, line_count + 1) * 30
            
            # 設置行高
            ws.row_dimensions[row].height = row_height
            print(f"第{row}行包含{line_count}行文本，設置行高為{row_height}磅 (相當於{max(2, line_count + 1)}行)")
        
        # 特別設置第1行(標題行)的高度為27
        ws.row_dimensions[1].height = 27
        print(f"已將第1行(標題行)的高度設置為27")
        
        # 设置A列单元格格式 - 对应VBA的设置A列格式部分
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            # 设置水平居中对齐，不自动换行
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False
            )
            # 设置字体大小为18
            new_font = Font(
                name="微軟正黑體",
                size=18,
                bold=True
            )
            cell.font = new_font
        
        # 设置B1单元格为水平和垂直居中，允许文本换行 - 对应VBA中对B1的设置
        b1 = ws.cell(row=1, column=2)
        b1.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # 设置页面布局为纵向 - 对应VBA的Orientation = xlPortrait
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        
        # 设置页眉 - 对应VBA的设置页眉部分
        current_date = datetime.now().strftime("%Y/%m/%d")
        ws.oddHeader.left.text = current_date
        ws.oddHeader.left.font = "微軟正黑體,粗體"
        ws.oddHeader.left.size = 20
        
        ws.oddHeader.center.text = "黑貓宅配(第一批)"
        ws.oddHeader.center.font = "微軟正黑體,粗體"
        ws.oddHeader.center.size = 20
        
        ws.oddHeader.right.text = "第 &P 頁，共 &N 頁"
        ws.oddHeader.right.font = "微軟正黑體,粗體"
        ws.oddHeader.right.size = 20
        
        # 设置页面边距（单位：英寸）- 对应VBA最后的边距设置
        ws.page_margins.left = 0.3937  # 约1厘米
        ws.page_margins.right = 0.3937  # 约1厘米
        ws.page_margins.top = 0.9843  # 约2.5厘米
        ws.page_margins.bottom = 0.3937  # 约1厘米
        ws.page_margins.header = 0.5118  # 约1.3厘米
        ws.page_margins.footer = 0.3937  # 约1厘米
        

        # 设置页面缩放为适应宽度 - 对应VBA的FitToPagesWide = 1, FitToPagesTall = 0
        # ws.page_setup.zoom = None  # 禁用按百分比縮放，必須設為None才能啟用fit_to_width
        # 1) 正確地設定「列印縮放」── fitToWidth / fitToHeight
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1  # 一頁寬
        ws.page_setup.fitToHeight = 0  # 高度不限（自動調整）

        # 2) 啟用 <pageSetUpPr fitToPage="1"/>  
        # 不然即使上面兩行，Excel 也會停留在「Adjust to XX%」模式
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # # （可選）關閉百分比縮放，確保「Adjust to」不會被啟用
        ws.page_setup.zoom = False

        # # 告訴 Excel 用 Page Layout 視圖
        # ws.sheet_view.view = "pageLayout"
        
        # 保存文件
        try:
            wb.save(temp_output)
            
            # 如果是临时文件，尝试重命名为最终输出文件
            if temp_output != output_file:
                try:
                    # 如果输出文件存在，先尝试删除
                    if os.path.exists(output_file):
                        os.remove(output_file)
                    # 重命名临时文件为最终输出文件
                    os.rename(temp_output, output_file)
                    print(f"处理完成，已保存到: {output_file}")
                    final_excel_path = output_file
                except Exception as e:
                    print(f"无法重命名为最终输出文件，输出已保存为临时文件: {temp_output}")
                    print(f"错误原因: {str(e)}")
                    final_excel_path = temp_output
            else:
                print(f"处理完成，已保存到: {output_file}")
                final_excel_path = output_file
                
        except PermissionError:
            # 如果出现权限错误，尝试保存到用户桌面
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            desktop_output = os.path.join(desktop, os.path.basename(output_file))
            wb.save(desktop_output)
            print(f"由于权限问题无法保存到原路径，已将文件保存到桌面: {desktop_output}")
            final_excel_path = desktop_output
            
        print("请在Excel中打开查看结果！")
        
        # 清理臨時檔案
        if is_json_data and os.path.exists(input_file) and not save_intermediate:
            try:
                os.remove(input_file)
                print(f"已清理臨時檔案: {input_file}")
            except:
                print(f"無法清理臨時檔案: {input_file}")
        
        # 導出為PDF
        if export_pdf:
            # 生成PDF文件路徑
            pdf_path = os.path.splitext(final_excel_path)[0] + '.pdf'
            # 嘗試導出為PDF
            export_result = export_excel_to_pdf(final_excel_path, pdf_path)
            if export_result:
                print(f"已成功導出PDF文件: {pdf_path}")
            else:
                if CURRENT_OS == 'Linux':
                    print("Linux環境下請確保已安裝LibreOffice")
                    print("可使用命令: sudo apt-get install libreoffice")
                elif CURRENT_OS == 'Windows' and not win32com_available:
                    print("Windows環境下請安裝pywin32")
                    print("可使用命令: pip install pywin32")
                else:
                    print(f"PDF導出失敗，請手動轉換")
        
        return final_excel_path
        
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False

# 如果直接运行此文件
if __name__ == "__main__":
    # 检查命令行参数
    if len(sys.argv) > 1 and sys.argv[1].lower() == 'json':
        # 使用示例JSON数据
        json_data = '''[
            {
                "name": "黃萍",
                "products": [
                    {
                        "productName": "購物金",
                        "amount": "1"
                    },
                    {
                        "productName": "預購(過膝靴/長靴)歐規尺碼 36-42 0119（樣式：37）",
                        "amount": "1"
                    }
                ]
            },
            {
                "name": "廖素津",
                "products": [
                    {
                        "productName": "購物金",
                        "amount": "1"
                    },
                    {
                        "productName": "預購(過膝靴/長靴)歐規尺碼 36-42 0119（樣式：37）",
                        "amount": "1"
                    }
                ]
            },
            {
                "name": "李淑娟",
                "products": [
                    {
                        "productName": "(預購)-YJ00001-潮流內刷毛外套(灰/橘/白)（樣式：橘L）",
                        "amount": "1"
                    }
                ]
            }
        ]'''
        
        # 输出文件路径
        output_file = os.path.join('output', 'json_formatted_output.xlsx')
        
        # 添加命令行参数支持
        export_pdf = True  # 默认导出PDF
        if len(sys.argv) > 2 and sys.argv[2].lower() == 'nopdf':
            export_pdf = False  # 如果命令行包含nopdf参数，则不导出PDF
        
        # 执行转换功能
        result = format_shipping_document(json_data, output_file, export_pdf)
    else:
        # 定义输入和输出文件路径
        input_file = os.path.join('input','翰禹_客戶出貨單_2025-04-15T11_30_55.xlsx')
        
        # 修改输出文件路径到同一目录但使用不同文件名
        output_file = os.path.join('output', 'formatted_output.xlsx')
        
        # 添加命令行参数支持
        export_pdf = True  # 默认导出PDF
        if len(sys.argv) > 1 and sys.argv[1].lower() == 'nopdf':
            export_pdf = False  # 如果命令行包含nopdf参数，则不导出PDF
        
        # 执行转换功能
        result = format_shipping_document(input_file, output_file, export_pdf)
    
    if result:
        print(f"成功完成處理，輸出檔案路徑: {result}")
    else:
        print("處理失敗！") 