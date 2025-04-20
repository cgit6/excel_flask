import functions_framework
import tempfile
import os
import openpyxl
from flask import jsonify, make_response, render_template_string

@functions_framework.http
def process_excel(request):
    """HTTP Cloud Function，用于接收、处理和返回Excel文件。
    
    Args:
        request (flask.Request): 请求对象
        
    Returns:
        处理后的Excel文件或错误信息
    """
    # 如果是GET请求，则显示上传表单
    if request.method == 'GET':
        html = '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>Excel文件處理</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 0; padding: 20px; line-height: 1.6; }
                .container { max-width: 800px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }
                h1 { color: #333; }
                .form-group { margin-bottom: 15px; }
                label { display: block; margin-bottom: 5px; font-weight: bold; }
                .btn { background-color: #4CAF50; color: white; padding: 10px 15px; border: none; cursor: pointer; }
                .btn:hover { background-color: #45a049; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Excel文件處理</h1>
                <p>請選擇要上傳的Excel文件 (.xlsx 或 .xls)</p>
                
                <form action="/process_excel" method="post" enctype="multipart/form-data">
                    <div class="form-group">
                        <label for="file">選擇文件:</label>
                        <input type="file" id="file" name="file" accept=".xlsx,.xls" required>
                    </div>
                    <button type="submit" class="btn">上傳並處理</button>
                </form>
            </div>
        </body>
        </html>
        '''
        return render_template_string(html)
    
    # 检查请求方法
    if request.method != 'POST':
        return jsonify({"error": "只支持POST请求"}), 405
    
    # 检查是否有文件上传
    if 'file' not in request.files:
        return jsonify({"error": "未找到上传的文件"}), 400
    
    uploaded_file = request.files['file']
    
    # 检查文件名是否为空
    if uploaded_file.filename == '':
        return jsonify({"error": "文件名为空"}), 400
    
    # 检查文件类型
    if not uploaded_file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持Excel文件(.xlsx或.xls)"}), 400
    
    try:
        # 创建临时文件来保存上传的Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
            temp_input_path = temp_input.name
            uploaded_file.save(temp_input_path)
        
        # 创建临时文件用于保存处理后的Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_output:
            temp_output_path = temp_output.name
        
        # 处理Excel文件
        process_result = process_excel_file(temp_input_path, temp_output_path)
        
        if not process_result:
            return jsonify({"error": "处理Excel文件时出错"}), 500
        
        # 读取处理后的文件并创建响应
        with open(temp_output_path, 'rb') as f:
            output_data = f.read()
        
        # 构建响应
        response = make_response(output_data)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=processed_{uploaded_file.filename}'
        
        # 清理临时文件
        try:
            os.unlink(temp_input_path)
            os.unlink(temp_output_path)
        except Exception as e:
            print(f"清理临时文件时出错: {str(e)}")
        
        return response
        
    except Exception as e:
        return jsonify({"error": f"处理请求时出错: {str(e)}"}), 500

def process_excel_file(input_path, output_path):
    """处理Excel文件的函数
    
    Args:
        input_path: 输入Excel文件路径
        output_path: 输出Excel文件路径
        
    Returns:
        bool: 处理是否成功
    """
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        # 在这里添加您的Excel处理逻辑
        # 例如：为第一行添加粗体
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            
        # 示例：在A1单元格添加标题
        ws['A1'] = ws['A1'].value or "处理后的数据"
        
        # 示例：修改所有单元格中的文本
        for row in ws.iter_rows(min_row=2):  # 从第二行开始
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.upper()  # 将文本转换为大写
        
        # 保存处理后的工作簿
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"处理Excel文件时出错: {str(e)}")
        return False

# 如果您想在本地测试此函数，可以使用以下代码
if __name__ == "__main__":
    # 注意：这仅用于本地测试，部署时会被忽略
    from flask import Flask, request
    
    app = Flask(__name__)
    
    @app.route('/', methods=['GET'])
    def index():
        return process_excel(request)
    
    @app.route('/process_excel', methods=['GET', 'POST'])
    def test_function():
        return process_excel(request)
    
    app.run(host='0.0.0.0', port=8080, debug=True)
